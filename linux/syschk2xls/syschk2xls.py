import os
import sys
import datetime
import subprocess
import logging
import glob
# import platform
# import configparser
import xlsxwriter
from openpyxl import load_workbook

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s = %(levelname)s - %(message)s')

def get_latest_config_file(prefix):
    files = glob.glob(f"{prefix}*.xlsx")
    if not files:
        return None
    return max(files, key=os.path.getctime)

def read_config_file(filename):
    config = {}
    wb = load_workbook(filename, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        if any(cell.value == 'IT구성정보명' for cell in row):
            header_row = row
            break

    if header_row is None:
        print("Error: 'IT구성정보명' 열을 찾을 수 없습니다.")
        return config

    headers = [cell.value for cell in header_row]
    id_column = headers.index('IT구성정보명')

    for row in ws.iter_rows(min_row=header_row[0].row + 1, values_only=True):
        if row[id_column]:
            row_data = dict(zip(headers, row))
            config[row[id_column]] = row_data

    return config

def get_distro():
    if os.path.exists('/etc/redhat-release'):
        return 'redhat'
    elif os.path.exists('/etc/debian_version'):
        return 'debian'
    else:
        return 'unknown'

def run_command(command, use_sudo=False):
    try:
        if use_sudo:
            full_command = f"sudo -E {command}"
        else:
            full_command = command
        logging.debug(f"Executing command: {full_command}")
        result = subprocess.run(full_command, shell=True, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        logging.debug(f"Command output: {result.stdout}")
        return result.stdout
    except subprocess.CalledProcessError as e:
        logging.error(f"Command failed: {e.stderr}")
        return f"Error: {e.stderr}"

def load_styles(filename):
    styles = {}
    with open(filename, 'r', encoding='utf-8') as f:
        exec(f.read(), globals(), styles)
    return styles

def load_commands(filename):
    with open(filename, 'r') as f:
        return f.read().splitlines()

def create_excel_report(config, results, styles):
    now = datetime.datetime.now()
    filename = f"system_diagnostics_report_{now.strftime('%Y%m%d-%H%M')}.xlsx"
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet("System Check Results")

    # Apply styles
    formatted_styles = {name: workbook.add_format(style) for name, style in styles.items()}

    # Write headers
    worksheet.write(0, 0, "시스템 점검 보고서", formatted_styles['title1_format'])
    worksheet.write(1, 0, f"점검일시: {now.strftime('%Y-%m-%d %H:%M')}", formatted_styles['title2_format'])

    headers = ["항목", "명령어", "결과", "상태"]
    for col, header in enumerate(headers):
        worksheet.write(2, col, header, formatted_styles['header1_format'])

    # Write results
    row = 3
    for item, data in results.items():
        worksheet.write(row, 0, item, formatted_styles['string1_format'])
        worksheet.write(row, 1, data['command'], formatted_styles['string2_format'])
        worksheet.write(row, 2, data['output'], formatted_styles['string1_format'])
        worksheet.write(row, 3, data['status'], formatted_styles['stat1_format'])
        row += 1

    # Auto-fit columns
    for col in range(len(headers)):
        worksheet.set_column(col, col, 20)

    workbook.close()
    print(f"Report generated: {filename}")

def main():
    distro = get_distro()
    print(f"Detected Linux distribution: {distro}")

    config_file = get_latest_config_file("구성관리조회")
    if not config_file:
        print("Error: No configuration file found.")
        sys.exit(1)

    config = read_config_file(config_file)
    if not config:
        print("Error: 구성 정보를 읽을 수 없습니다. 프로그램을 종료합니다.")
        sys.exit(1)

    styles = load_styles('template_styles.conf')
    commands = load_commands('template_command.conf')

    results = {}
    for command in commands:
        if command.startswith('#') or command.strip() == '':
            continue
        if '=' in command:
            variable, cmd = command.split('=', 1)
            cmd = cmd.strip()
            use_sudo = True
            output = run_command(cmd, use_sudo=use_sudo)
            results[variable.strip()] = {
                'command': cmd,
                'output': output,
                'status': 'OK' if output and not output.startswith('Error:') else 'Error'
            }
        elif command.startswith('./'):
            script_name = os.path.basename(command)
            output = run_command(f"bash {command}")
            results[script_name] = {
                'command': command,
                'output': output,
                'status': 'OK' if output and not output.startswith('Error:') else 'Error'
            }

    create_excel_report(config, results, styles)

if __name__ == "__main__":
    main()
